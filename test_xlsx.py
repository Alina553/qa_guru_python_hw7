from openpyxl import load_workbook
import os
#4TODO оформить в тест, добавить ассерты и использовать универсальный путь

FILES_PATH = os.path.abspath(__file__)
ROOT_PATH = os.path.dirname(FILES_PATH)
RESOURCE_PATH = os.path.join(ROOT_PATH, 'resources')

def test_xlsx():
    workbook = load_workbook(os.path.join(RESOURCE_PATH,'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)

    assert (sheet.cell(row=3, column=2).value) == 'Mara'

